import os
import sys
import random
try:
    import openpyxl
except ImportError:
    import pip
    pip.main(['install', "openpyxl"])
import warnings
warnings.filterwarnings("ignore")

'''
Glory to Knattholmen! #Arstoztka
'''

#cell  defs
def cell(attribute):
    cell = {
                "Land": ws["D4"],
                "Land_eng": ws["E4"],
                "Etternavn": ws["D7"],
                "Fornavn": ws["F7"],
                "Fødeselsdato": ws["D10"],
                "Fødested": ws["F10"],
                "Nasjonalitet": ws["D13"],
                "Fødselsnummer": ws["F13"]
    }
    return cell[attribute]

genders = ["Boy", "Girl"]

def make_bday():
    '''
    Because a func is cleaner than doing this elswhere.
    '''
    zero_digits = ["0" + str(i) for i in range(10)]
    bday = []
    bday.append(str(random.choice(zero_digits + list(range(10, 29)))))
    bday.append(str(random.choice(zero_digits + list(range(10, 13)))))
    bday.append(str(random.choice(list(range(90, 100)) + zero_digits[:6]))) #fiks flere årstall 2000-2005
    return bday

def make_prs_num(bday, gender):
    '''
    The Norwegian algorythm for creating unique personal numbers (ID)
    http://www.matematikk.org/artikkel.html?tid=64296
    '''

    prs_num = ''
    if int(bday[-1]) <= 1999:
        prs_num += str(random.randint(0, 4)) + str(random.randint(0, 9)) #00-50 exclusive
    elif int(bday[-1]) >= 2000:
        prs_num += str(random.ranint(50, 100))
    if gender == "Girl":
        prs_num += random.choice(["2", "4", "6", "8"])
    elif gender == "Boy":
        prs_num += random.choice(["1", "3", "5", "7", "9"])

    #pardon the mess
    formula_controldigit_1 = [3, 7, 6, 1, 8, 9, 4, 5, 2]
    #V1=3D1+7D2+6M1+M2+8å1+9å2+4I1+5I2+2I3 #Algorythm for first control digit
    formula_controldigit_2 = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    #V2=5D1+4D2+3M1+2M2+7å1+6å2+5I1+4I2+3I3+2K1 #Algorythm for second control digit

    temp_control_digit = 0
    for index, i in enumerate(''.join(bday) + prs_num):
        temp_control_digit += (int(i) * formula_controldigit_1[index])
    if temp_control_digit % 11 == 0:
        prs_num += str(0)
    else:
        prs_num += str(11 - (temp_control_digit % 11))

    temp_control_digit = 0
    for index, i in enumerate(''.join(bday) + prs_num):
        temp_control_digit += (int(i) * formula_controldigit_2[index])
    #prs_num += str(temp_control_digit % 11)
    if temp_control_digit % 11 == 0:
        prs_num += str(0)
    else:
        prs_num += str(11 - (temp_control_digit % 11))

    return prs_num

country_data = {
                "Country 1" :
                        {
                        "Boy name": ["Boy1", "Boy2", "Boy3"],
                        "Girl name": ["Girl1", "Girl2", "Girl3"],
                        "Family name": ["Family1", "Family2", "Family3"],
                        "Birthplace": ["City1", "City2", "City3"],
                        "Nor_eng_name": ["Land", "Country"],
                        "Nationality": ["Landsk", "Country-ian"],
                        "Abbreviation": "ABC"
                        },
                "Country 2" :
                        {
                        "Boy name": ["Boy1", "Boy2", "Boy3"],
                        "Girl name": ["Girl1", "Girl2", "Girl3"],
                        "Family name": ["Family1", "Family2", "Family3"],
                        "Birthplace": ["City1", "City2", "City3"],
                        "Nor_eng_name": ["Land", "Country"],
                        "Nationality": ["Landsk", "Country-ian"],
                        "Abbreviation": "ABC"
                        },
                "Country 3" :
                        {
                        "Boy name": ["Boy1", "Boy2", "Boy3"],
                        "Girl name": ["Girl1", "Girl2", "Girl3"],
                        "Family name": ["Family1", "Family2", "Family3"],
                        "Birthplace": ["City1", "City2", "City3"],
                        "Nor_eng_name": ["Land", "Country"],
                        "Nationality": ["Landsk", "Country-ian"],
                        "Abbreviation": "ABC"
                        },
                "Country 4" :
                        {
                        "Boy name": ["Boy1", "Boy2", "Boy3"],
                        "Girl name": ["Girl1", "Girl2", "Girl3"],
                        "Family name": ["Family1", "Family2", "Family3"],
                        "Birthplace": ["City1", "City2", "City3"],
                        "Nor_eng_name": ["Land", "Country"],
                        "Nationality": ["Landsk", "Country-ian"],
                        "Abbreviation": "ABC"
                        },
                "Country 5" :
                        {
                        "Boy_name": ["Boy1", "Boy2", "Boy3"],
                        "Girl_name": ["Girl1", "Girl2", "Girl3"],
                        "Family_name": ["Family1", "Family2", "Family3"],
                        "Birthplace": ["City1", "City2", "City3"],
                        "Nor_eng_name": ["Land", "Country"],
                        "Nationality": ["Landsk", "Country-ian/ish/an"],
                        "Abbreviation": "ABC"
                        }
}

def make_passport(country, gender, n, bday=make_bday()):
    global country_data

   # country_data = country_data[country]

    wb = openpyxl.load_workbook("Pass_mal.xlsx")
    #wb = openpyxl.load_workbook("Pass_mal.xlsx")
    ws = wb.get_active_sheet()

    ws["D4"] = country_data[country]["Nor_eng_name"][0].upper()
    ws["E4"] = country_data[country]["Nor_eng_name"][1].upper()
    ws["D7"] = random.choice(country_data[country]["Family name"])
    ws["F7"] = random.choice(country_data[country][gender + " name"])
    ws["D10"] = '.'.join(bday)
    ws["F10"] = "{}, {}".format(random.choice(  country_data[country]["Birthplace"]).upper(),
                                                country_data[country]["Abbreviation"].upper()) #.upper
    ws["D13"] = ' '.join(country_data[country]["Nationality"])
    ws["F13"] = "{}{}".format( ''.join(bday),
                                make_prs_num(bday, gender)) #fiks blokkbokstaver

    wb.save("TestPassport.xlsx")

country_list = [country for country in country_data.keys()]
for country in country_list:
    for gender in ["Boy", "Girl"]:
        for i in range(10):
            pass
            #make_passport(country, gender, i)
    
make_passport("Country 1", "Boy", str(1))
print("DONE")
